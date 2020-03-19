
import os
import xml.etree.ElementTree as ET
import openpyxl
import csv

'''
    Result tree is a hierarchical set of results consisting of result entries
    Result root element is the only elemement in the tree with the type 'result'
    Results data is contained in files in the file system
    Results file name attributes in results are all relative to the root result directory
'''

RESULT_TYPE_RESULT = 'result'
RESULT_TYPE_SUITE = 'suite'
RESULT_TYPE_TEST = 'test'
RESULT_TYPE_SCRIPT = 'script'
RESULT_TYPE_FILE = 'file'

type_ext = {RESULT_TYPE_SUITE: '.ste',
            RESULT_TYPE_TEST: '.tst',
            RESULT_TYPE_SCRIPT: '.py'}

RESULT_RUNNING = 'Running'
RESULT_STOPPED = 'Stopped'
RESULT_COMPLETE = 'Complete'
RESULT_PASS = 'Pass'
RESULT_FAIL = 'Fail'

PARAM_TYPE_STR = 'string'
PARAM_TYPE_INT = 'int'
PARAM_TYPE_FLOAT = 'float'
PARAM_TYPE_BOOL = 'bool'

param_types = {'int': int, 'float': float, 'string': str, 'bool': bool,
               int: 'int', float: 'float', str: 'string', bool: 'bool'}

RESULT_TAG = 'result'
RESULT_ATTR_NAME = 'name'
RESULT_ATTR_TYPE = 'type'
RESULT_ATTR_STATUS = 'status'
RESULT_ATTR_FILENAME = 'filename'
RESULT_PARAMS = 'params'
RESULT_PARAM = 'param'
RESULT_PARAM_ATTR_NAME = 'name'
RESULT_PARAM_ATTR_TYPE = 'type'
RESULT_RESULTS = 'results'


class ResultError(Exception):
    pass


class ResultWorkbook(object):

    def __init__(self, filename=None):
        self.wb = openpyxl.Workbook()
        self.filename = None

        # remove initial sheet that is added at creation
        self.wb.remove_sheet(self.wb.active)

    def add_chart(self, ws, params=None):
        # get fieldnames in first row of worksheet
        colors = ['blue', 'green', 'purple', 'orange', 'red']
        color_idx = 0
        point_names = []
        for c in ws.rows[0]:
            point_names.append(c.value)

        x_points = []
        y_points = []
        y2_points = []
        if params is not None:
            points = params.get('plot.x.points')
            if points is not None:
                x_points = [x.strip() for x in points.split(',')]
            points = params.get('plot.y.points')
            if points is not None:
                y_points = [x.strip() for x in points.split(',')]
            points = params.get('plot.y2.points')
            if points is not None:
                y2_points = [x.strip() for x in points.split(',')]

        chart = openpyxl.chart.ScatterChart(scatterStyle='line')
        chart.title = params.get('plot.title', None)
        chart.style = 13
        chart.x_axis.title = params.get('plot.x.title', '')
        chart.y_axis.title = params.get('plot.y.title', '')

        x_values = None
        if len(x_points) > 0:
            # only support one x point for now
            name = x_points[0]
            try:
                col = point_names.index(name) + 1
                print('x: %s %s' % (col, ws.max_row))
                x_values = openpyxl.chart.Reference(ws, min_col=col, min_row=2, max_row=ws.max_row)
            except ValueError:
                pass

        if len(y_points) > 0:
            for name in y_points:
                try:
                    col = point_names.index(name) + 1
                    values = openpyxl.chart.Reference(ws, min_col=col, min_row=2, max_row=ws.max_row)
                    series = openpyxl.chart.Series(values, x_values, title=name)

                    # lineProp = drawing.line.LineProperties(prstDash='dash')
                    lineProp = openpyxl.drawing.line.LineProperties(
                        solidFill = openpyxl.drawing.colors.ColorChoice(prstClr=colors[color_idx]))
                    color_idx += 1
                    series.graphicalProperties.line = lineProp
                    series.graphicalProperties.line.width = 20000 # width in EMUs
                    chart.series.append(series)

                except ValueError:
                    pass

        if len(y2_points) > 0:
            for name in y2_points:
                try:
                    col = point_names.index(name) + 1
                    values = openpyxl.chart.Reference(ws, min_col=col, min_row=2, max_row=ws.max_row)
                    series = openpyxl.chart.Series(values, x_values, title=name)

                    # lineProp = drawing.line.LineProperties(prstDash='dash')
                    lineProp = openpyxl.drawing.line.LineProperties(
                        solidFill = openpyxl.drawing.colors.ColorChoice(prstClr=colors[color_idx]))
                    color_idx += 1
                    series.graphicalProperties.line = lineProp
                    series.graphicalProperties.line.width = 20000 # width in EMUs
                    chart2 = openpyxl.chart.ScatterChart(scatterStyle='line')
                    chart2.style = 13
                    # chart.y_axis.title = params.get('plot.y.title', '')
                    chart2.series.append(series)
                    chart2.y_axis.axId = 200
                    chart2.y_axis.title = params.get('plot.y2.title', '')
                    chart.y_axis.crosses = "max"
                    chart += chart2

                except ValueError:
                    pass

        idx = self.wb.sheetnames.index(ws.title) - 1
        if idx < 0:
            idx = 0
        cs = self.wb.create_chartsheet(title=params.get('plot.title', None), index=idx)
        cs.add_chart(chart)

    def add_csv_file(self, filename, title, relative_value_names=None, params=None):
        line = 1
        ws = self.wb.create_sheet(title=title)
        f = None
        relative_value_index = []
        relative_value_start = []
        if relative_value_names is None:
            relative_value_names = []
        try:
            f = open(filename)
            reader = csv.reader(f, skipinitialspace=True)
            for row in reader:
                for i in range(len(row)):
                    try:
                        row[i] = float(row[i])
                    except ValueError:
                        pass
                # find fields to be treated as relative value
                if line == 1:
                    line += 1
                    if relative_value_names is not None:
                        for name in relative_value_names:
                            try:
                                index = row.index(name)
                                relative_value_index.append(index)
                            except ValueError:
                                pass
                # get initial value for relative value fields
                elif line == 2:
                    line += 1
                    for index in relative_value_index:
                        relative_value_start.append(row[index])
                        row[index] = 0
                else:
                    for index in relative_value_index:
                        row[index] = row[index] - relative_value_start[index]
                ws.append(row)

            if title[-4:] == '.csv':
                chart_title = title[:-4]
            else:
                chart_title = title + '_chart'

            self.add_chart(ws, params=params)
            '''
            self.add_chart(ws, params={'plot.title': chart_title,
                                       'plot.x.title': 'Time (secs)',
                                       'plot.x.points': 'TIME',
                                       'plot.y.points': 'AC_VRMS_1',
                                       'plot.y.title': 'Voltage (V)',
                                       'plot.y2.points': 'AC_IRMS_1',
                                       'plot.y2.title': 'Current (A)'})
            '''
        except Exception as e:
            raise
        finally:
            if f:
                f.close()

    def save(self, filename=None):
        if filename:
            self.filename = filename
        self.wb.save(self.filename)


class Result(object):

    def __init__(self, name=None, type=None, status=None, filename=None, params=None, result_path=None):
        self.name = name
        self.type = type
        self.status = status
        self.filename = filename
        self.params = []
        self.result_path = result_path
        self.ref = None
        self.results_index = 0
        if params is not None:
            self.params = params
        else:
            self.params = {}
        self.results = []

    def __str__(self):
        return self.to_str()

    def next_result(self):
        if self.results_index < len(self.results):
            result = self.results[self.results_index]
            self.results_index += 1
            return result

    def add_result(self, result):
        self.results.append(result)

    def file(self):
        return self.name + type_ext.get(self.type, '')

    def to_str(self, indent=''):
        s = '%sname = %s  type = %s  status = %s  filename = %s\n%s  params = %s\n%s  results = \n  ' % (
            indent, self.name, self.type, self.status, self.filename, indent, self.params, indent
        )
        indent += '  '
        for r in self.results:
            s += '%s' % (r.to_str(indent=indent))
        return s

    def from_xml(self, element=None, filename=None):
        if element is None and filename is not None:
            element = ET.ElementTree(file=filename).getroot()
            self.result_path, file = os.path.split(filename)
        if element is None:
            raise ResultError('No xml document element')
        if element.tag != RESULT_TAG:
            raise ResultError('Unexpected result root element: %s' % (element.tag))
        self.name = element.attrib.get(RESULT_ATTR_NAME)
        self.type = element.attrib.get(RESULT_ATTR_TYPE)
        self.status = element.attrib.get(RESULT_ATTR_STATUS)
        self.filename = element.attrib.get(RESULT_ATTR_FILENAME)
        self.params = {}
        self.results = []
        if self.name is None:
            raise ResultError('Result name missing')

        for e in element.findall('*'):
            if e.tag == RESULT_PARAMS:
                for e_param in e.findall('*'):
                    if e_param.tag == RESULT_PARAM:
                        name = e_param.attrib.get(RESULT_PARAM_ATTR_NAME)
                        param_type = e_param.attrib.get(RESULT_PARAM_ATTR_TYPE)
                        if name:
                            vtype = param_types.get(param_type, str)
                            self.params[name] = vtype(e_param.text)
            elif e.tag == RESULT_RESULTS:
                for e_param in e.findall('*'):
                    if e_param.tag == RESULT_TAG:
                        result = Result(result_path=self.result_path)
                        self.results.append(result)
                        result.from_xml(e_param)

    def to_xml(self, parent=None, filename=None):
        attr = {}
        if self.name:
            attr[RESULT_ATTR_NAME] = self.name
        if self.type:
            attr[RESULT_ATTR_TYPE] = self.type
        if self.status:
            attr[RESULT_ATTR_STATUS] = self.status
        if self.filename:
            attr[RESULT_ATTR_FILENAME] = self.filename
        if parent is not None:
            e = ET.SubElement(parent, RESULT_TAG, attrib=attr)
        else:
            e = ET.Element(RESULT_TAG, attrib=attr)

        e_params = ET.SubElement(e, RESULT_PARAMS)

        params = sorted(self.params, key=self.params.get)
        for p in params:
            value_type = None
            value_str = None
            attr = {RESULT_PARAM_ATTR_NAME: p}
            value = self.params.get(p)
            if value is not None:
                value_type = param_types.get(type(value), PARAM_TYPE_STR)
                value_str = str(value)

            if value_type is not None:
                attr[RESULT_PARAM_ATTR_TYPE] = value_type

            e_param = ET.SubElement(e_params, RESULT_PARAM, attrib=attr)
            if value_str is not None:
                e_param.text = value_str

        e_results = ET.SubElement(e, RESULT_RESULTS)
        for r in self.results:
            r.to_xml(e_results)

        return e

    def to_xml_str(self, pretty_print=False):
        e = self.to_xml()

        if pretty_print:
            xml_indent(e)

        return ET.tostring(e)

    def to_xml_file(self, filename=None, pretty_print=True, replace_existing=True):
        xml = self.to_xml_str(pretty_print)
        if filename is None and self.filename is not None:
            filename = self.filename

        if filename is not None:
            if replace_existing is False and os.path.exists(filename):
                raise ResultError('File %s already exists' % (filename))
            f = open(filename, 'w')
            f.write(xml)
            f.close()
        else:
            print(xml)

    def to_xlsx(self, wb=None, filename=None):
        '''
        self.params={'plot.title': self.name,
                     'plot.x.title': 'Time (secs)',
                     'plot.x.points': 'TIME',
                     'plot.y.points': 'AC_VRMS_1',
                     'plot.y.title': 'Voltage (V)',
                     'plot.y2.points': 'AC_IRMS_1',
                     'plot.y2.title': 'Current (A)'}
        '''
        print('creating to_xlsx for %s' % filename)

        result_wb = wb
        if result_wb is None:
            result_wb = ResultWorkbook(filename=filename)
        if self.type == RESULT_TYPE_FILE:
            name, ext = os.path.splitext(self.filename)
            if ext == '.csv':
                result_wb.add_csv_file(os.path.join(self.result_path, self.filename), self.name,
                                       relative_value_names = ['TIME'], params=self.params)
        for result in self.results:
            result.to_xlsx(wb=result_wb)
        if wb is None:
            print('saving')
            result_wb.save(filename=filename)

""" Simple XML pretty print support function

"""
def xml_indent(elem, level=0):
    i = "\n" + level*"  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "  "
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
        for elem in elem:
            xml_indent(elem, level+1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i

if __name__ == "__main__":

    '''
    result = Result(name='Result', type='suite')
    result1 = Result(name='Result 1', type='test', status='complete')
    result1.results.append(Result(name='Result 1 Log', type='log', filename='log/file/name/1'))
    result2 = Result(name='Result 2', type='test', status='complete', params={'param 1': 'param 1 value'})
    result2.results.append(Result(name='Result 2 Log', type='log', filename='log/file/name/2'))
    result.results.append(result1)
    result.results.append(result2)

    xml_str = result.to_xml_str(pretty_print=True)
    print xml_str
    print result
    print '-------------------'
    result_xml = Result()
    root = ET.fromstring(xml_str)
    result_xml.from_xml(root)
    print result_xml
    '''
    result_xml = Result()
    result_xml.from_xml(filename='c:\\users\\bob\\downloads\\ul1741 sa_ucsd_6\\ul1741 sa 6\\results\\2017-06-28_13-24-40-977__HLVRT__LVRT_LV2\\2017-06-28_13-24-40-977__HLVRT__LVRT_LV2.rlt')
    result_xml.to_xlsx(filename='c:\\users\\bob\\downloads\\ul1741 sa_ucsd_6\\ul1741 sa 6\\results\\2017-06-28_13-24-40-977__HLVRT__LVRT_LV2\\2017-06-28_13-24-40-977__HLVRT__LVRT_LV2.xlsx')
    print(result_xml)

